"""Service for building and executing dynamic report queries."""

import logging
from typing import Any, Dict, List, Optional
from io import BytesIO

from repositories.saved_report_repository import (
    ReportQueryRequest,
    SavedReportCreate,
    SavedReportUpdate,
)
from repositories.report_repository import (
    get_entity_model,
    execute_report_query,
    fetch_polymorphic_entity,
    fetch_notes_for_entity,
    get_leads_for_pipeline,
    get_account_overview_counts,
    get_organizations_with_contacts,
    get_notes_for_activity_report,
    get_entity_name_for_note,
    get_lead_type_for_note,
    get_user_by_id,
    get_audit_models,
    get_audit_counts_by_table,
    get_recent_audit_activity,
)

# Re-export Pydantic models for controllers
__all__ = ["ReportQueryRequest", "SavedReportCreate", "SavedReportUpdate"]

logger = logging.getLogger(__name__)

ENTITY_FIELDS = {
    "organizations": [
        "id", "name", "website", "phone", "employee_count", "description",
        "founding_year", "headquarters_city", "headquarters_state", "headquarters_country",
        "company_type", "linkedin_url", "crunchbase_url", "created_at", "updated_at"
    ],
    "individuals": [
        "id", "first_name", "last_name", "email", "phone", "title", "department",
        "seniority_level", "linkedin_url", "twitter_url", "github_url", "bio",
        "is_decision_maker", "created_at", "updated_at"
    ],
    "contacts": [
        "id", "first_name", "last_name", "email", "phone", "title", "department",
        "role", "is_primary", "notes", "created_at", "updated_at"
    ],
    "leads": [
        "id", "title", "description", "source", "type", "created_at", "updated_at"
    ],
    "notes": [
        "id", "entity_type", "entity_id", "content", "created_at", "updated_at"
    ],
}

JOIN_FIELDS = {
    "organizations": {
        "account": ["account.name"],
        "employee_count_range": ["employee_count_range.label"],
        "funding_stage": ["funding_stage.name"],
        "revenue_range": ["revenue_range.label"],
    },
    "individuals": {
        "account": ["account.name"],
    },
    "contacts": {
        "organizations": ["organization.name"],
        "individuals": ["individual.first_name", "individual.last_name"],
    },
    "leads": {
        "account": ["account.name"],
    },
    "notes": {},
}

ENTITY_JOINS = {
    "notes": {
        "leads": {"entity_type": "Lead", "fields": ["lead.title", "lead.source", "lead.type"]},
        "organizations": {"entity_type": "Organization", "fields": ["organization.name", "organization.website"]},
        "individuals": {"entity_type": "Individual", "fields": ["individual.first_name", "individual.last_name", "individual.email"]},
        "contacts": {"entity_type": "Contact", "fields": ["contact.first_name", "contact.last_name", "contact.email"]},
    },
    "leads": {
        "notes": {"fields": ["note.content", "note.created_at"]},
    },
    "organizations": {
        "notes": {"fields": ["note.content", "note.created_at"]},
    },
    "individuals": {
        "notes": {"fields": ["note.content", "note.created_at"]},
    },
    "contacts": {
        "notes": {"fields": ["note.content", "note.created_at"]},
    },
}

# Python-side operators for filtering joined data in memory
PYTHON_OPERATORS = {
    "eq": lambda val, filter_val: val == filter_val,
    "neq": lambda val, filter_val: val != filter_val,
    "gt": lambda val, filter_val: val is not None and val > filter_val,
    "gte": lambda val, filter_val: val is not None and val >= filter_val,
    "lt": lambda val, filter_val: val is not None and val < filter_val,
    "lte": lambda val, filter_val: val is not None and val <= filter_val,
    "contains": lambda val, filter_val: val is not None and str(filter_val).lower() in str(val).lower(),
    "starts_with": lambda val, filter_val: val is not None and str(val).lower().startswith(str(filter_val).lower()),
    "is_null": lambda val, filter_val: val is None or val == "",
    "is_not_null": lambda val, filter_val: val is not None and val != "",
    "in": lambda val, filter_val: val in (filter_val if isinstance(filter_val, list) else [filter_val]),
}

ENTITY_TYPE_MAP = {
    "leads": "Lead",
    "organizations": "Organization",
    "individuals": "Individual",
    "contacts": "Contact",
}


def get_available_fields(entity: str) -> Dict[str, Any]:
    """Get available fields for an entity including join fields."""
    base_fields = ENTITY_FIELDS.get(entity, [])
    joins = JOIN_FIELDS.get(entity, {})
    join_field_list = []
    for join_fields in joins.values():
        join_field_list.extend(join_fields)

    entity_joins = ENTITY_JOINS.get(entity, {})
    available_joins = []
    for join_name, join_config in entity_joins.items():
        available_joins.append({
            "name": join_name,
            "fields": join_config["fields"],
        })
        join_field_list.extend(join_config["fields"])

    return {
        "base": base_fields,
        "joins": join_field_list,
        "available_joins": available_joins,
    }


def _format_value(val: Any) -> Any:
    """Format a value, converting datetime objects to ISO format."""
    if hasattr(val, "isoformat"):
        return val.isoformat()
    return val


def _get_direct_field(row, col: str) -> Any:
    """Get a direct field value from a row."""
    if not hasattr(row, col):
        return None
    return _format_value(getattr(row, col))


def _get_joined_field(joined_rows: Dict[str, Any], rel_name: str, field_name: str) -> Any:
    """Get a field value from joined rows."""
    joined_row = joined_rows.get(rel_name)
    if not joined_row or not hasattr(joined_row, field_name):
        return None
    return _format_value(getattr(joined_row, field_name))


def _get_relationship_field(row, rel_name: str, field_name: str) -> Any:
    """Get a field value from a model relationship."""
    rel_obj = getattr(row, rel_name, None)
    if not rel_obj:
        return None
    return _format_value(getattr(rel_obj, field_name, None))


def _extract_field_value(row, col: str, joined_rows: Dict[str, Any]):
    """Extract a field value from row or joined data."""
    if "." not in col:
        return _get_direct_field(row, col)

    rel_name, field_name = col.split(".", 1)

    joined_val = _get_joined_field(joined_rows, rel_name, field_name)
    if joined_val is not None:
        return joined_val
    return _get_relationship_field(row, rel_name, field_name)


def _row_passes_join_filters(row_dict: Dict[str, Any], join_filters: List[Dict[str, Any]]) -> bool:
    """Check if a row passes all join field filters."""
    valid_filters = [f for f in join_filters if f.get("field") and f.get("operator")]
    for f in valid_filters:
        field_value = row_dict.get(f["field"])
        op_func = PYTHON_OPERATORS.get(f["operator"])
        if op_func and not op_func(field_value, f.get("value")):
            return False
    return True


async def _fetch_polymorphic_join(row, join_config, join_name) -> tuple:
    """Fetch polymorphic join for notes -> entity."""
    if row.entity_type != join_config.get("entity_type"):
        return None, None
    entity = await fetch_polymorphic_entity(join_config["entity_type"], row.entity_id)
    singular_key = join_name.rstrip("s") if join_name.endswith("s") else join_name
    return singular_key, entity


async def _fetch_notes_join(row, primary_entity: str, tenant_id: int) -> tuple:
    """Fetch notes join for entity -> notes."""
    entity_type = ENTITY_TYPE_MAP.get(primary_entity)
    if not entity_type:
        return None, None
    note = await fetch_notes_for_entity(tenant_id, entity_type, row.id)
    return "note", note


async def _fetch_single_join(
    row, join_name: str, join_config: Dict, primary_entity: str, tenant_id: int
) -> tuple:
    """Fetch a single join based on entity type."""
    if primary_entity == "notes" and join_config.get("entity_type"):
        return await _fetch_polymorphic_join(row, join_config, join_name)

    if join_name == "notes":
        return await _fetch_notes_join(row, primary_entity, tenant_id)

    return None, None


async def _fetch_joined_rows(
    row, joins: List[str], primary_entity: str, tenant_id: int
) -> Dict[str, Any]:
    """Fetch all joined data for a row."""
    joined_rows = {}
    entity_joins = ENTITY_JOINS.get(primary_entity, {})

    for join_name in joins:
        join_config = entity_joins.get(join_name)
        if not join_config:
            continue
        key, value = await _fetch_single_join(row, join_name, join_config, primary_entity, tenant_id)
        if key:
            joined_rows[key] = value

    return joined_rows


async def execute_custom_query(
    tenant_id: int,
    primary_entity: str,
    columns: List[str],
    joins: Optional[List[str]] = None,
    filters: Optional[List[Dict[str, Any]]] = None,
    group_by: Optional[List[str]] = None,
    aggregations: Optional[List[Dict[str, Any]]] = None,
    limit: int = 100,
    offset: int = 0,
    project_id: Optional[int] = None,
) -> Dict[str, Any]:
    """Execute a custom report query and return results."""
    model = get_entity_model(primary_entity)
    if not model:
        return {"error": f"Unknown entity: {primary_entity}", "data": [], "total": 0}

    joins = joins or []

    rows, total = await execute_report_query(
        tenant_id=tenant_id,
        primary_entity=primary_entity,
        filters=filters,
        limit=limit,
        offset=offset,
        project_id=project_id,
    )

    # Fetch joined data for each row
    data = []
    for row in rows:
        joined_rows = await _fetch_joined_rows(row, joins, primary_entity, tenant_id)
        row_dict = {col: _extract_field_value(row, col, joined_rows) for col in columns}

        # Apply join field filters (post-fetch filtering)
        join_filters = [f for f in (filters or []) if "." in f.get("field", "")]
        if _row_passes_join_filters(row_dict, join_filters):
            data.append(row_dict)

    return {"data": data, "total": len(data), "limit": limit, "offset": offset}


async def get_lead_pipeline_report(
    tenant_id: int,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    project_id: Optional[int] = None
) -> Dict[str, Any]:
    """Generate lead pipeline canned report."""
    leads = await get_leads_for_pipeline(tenant_id, date_from, date_to, project_id)

    type_counts = {}
    source_counts = {}
    total = len(leads)

    for lead in leads:
        lead_type = lead.type or "Unknown"
        type_counts[lead_type] = type_counts.get(lead_type, 0) + 1

        source = lead.source or "Unknown"
        source_counts[source] = source_counts.get(source, 0) + 1

    return {
        "total_leads": total,
        "by_type": type_counts,
        "by_source": source_counts,
    }


async def get_account_overview_report(tenant_id: int) -> Dict[str, Any]:
    """Generate account overview canned report."""
    return await get_account_overview_counts(tenant_id)


async def get_contact_coverage_report(tenant_id: int) -> Dict[str, Any]:
    """Generate contact coverage canned report."""
    orgs = await get_organizations_with_contacts(tenant_id)

    total_orgs = len(orgs)
    total_contacts = 0
    orgs_with_zero_contacts = 0
    decision_makers = 0
    total_contact_count = 0

    coverage_data = []
    for org in orgs:
        contacts = org.account.contacts if org.account else []
        contact_count = len(contacts)
        total_contacts += contact_count
        if contact_count == 0:
            orgs_with_zero_contacts += 1
        for contact in contacts:
            total_contact_count += 1
            if getattr(contact, "is_decision_maker", False):
                decision_makers += 1
        coverage_data.append({
            "organization_id": org.id,
            "organization_name": org.name,
            "contact_count": contact_count,
        })

    avg_contacts = total_contacts / total_orgs if total_orgs > 0 else 0
    decision_maker_ratio = decision_makers / total_contact_count if total_contact_count > 0 else 0

    return {
        "total_organizations": total_orgs,
        "total_contacts": total_contacts,
        "average_contacts_per_org": round(avg_contacts, 2),
        "organizations_with_zero_contacts": orgs_with_zero_contacts,
        "decision_maker_count": decision_makers,
        "decision_maker_ratio": round(decision_maker_ratio, 4),
        "coverage_by_org": sorted(coverage_data, key=lambda x: x["contact_count"], reverse=True)[:20],
    }


async def get_notes_activity_report(tenant_id: int, project_id: Optional[int] = None) -> Dict[str, Any]:
    """Generate notes activity canned report."""
    notes = await get_notes_for_activity_report(tenant_id, project_id)

    total_notes = len(notes)

    by_entity_type = {}
    for note in notes:
        entity_type = note.entity_type or "Unknown"
        by_entity_type[entity_type] = by_entity_type.get(entity_type, 0) + 1

    entities_with_notes = {}
    entity_ids_seen = {}
    for note in notes:
        key = f"{note.entity_type}:{note.entity_id}"
        if key not in entity_ids_seen:
            entity_ids_seen[key] = True
            entity_type = note.entity_type or "Unknown"
            entities_with_notes[entity_type] = entities_with_notes.get(entity_type, 0) + 1

    recent_notes = []
    for note in notes[:10]:
        entity_name = await get_entity_name_for_note(note.entity_type, note.entity_id)
        lead_type = await get_lead_type_for_note(note.entity_type, note.entity_id)
        recent_notes.append({
            "id": note.id,
            "entity_type": note.entity_type,
            "entity_id": note.entity_id,
            "entity_name": entity_name,
            "lead_type": lead_type,
            "content": note.content[:100] + "..." if note.content and len(note.content) > 100 else note.content,
            "created_at": note.created_at.isoformat() if note.created_at else None,
        })

    return {
        "total_notes": total_notes,
        "by_entity_type": by_entity_type,
        "entities_with_notes": entities_with_notes,
        "recent_notes": recent_notes,
    }


async def get_user_audit_report(tenant_id: int, user_id: Optional[int] = None) -> Dict[str, Any]:
    """Generate user audit report showing records created/updated by users."""
    user_info = None
    if user_id:
        user_info = await get_user_by_id(user_id)

    by_table, total_created, total_updated = await get_audit_counts_by_table(tenant_id, user_id)

    recent_activity = await get_recent_audit_activity(tenant_id, user_id, limit_per_table=5)
    recent_activity.sort(key=lambda x: x["updated_at"] or "", reverse=True)
    recent_activity = recent_activity[:20]

    return {
        "user": user_info,
        "total_created": total_created,
        "total_updated": total_updated,
        "by_table": by_table,
        "recent_activity": recent_activity,
    }


def generate_excel_bytes(data: List[Dict[str, Any]], columns: List[str]) -> bytes:
    """Generate Excel file bytes from report data."""
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
